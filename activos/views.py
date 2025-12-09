from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView, DetailView, UpdateView, DeleteView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
import csv
import json
from openpyxl import Workbook
from django.db import models
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import get_user_model
from datetime import timedelta
from .models import Activo, Tranzabilidad, Historial, Zona, Categoria, Marca
from .forms import ActivoForm

@login_required
def dashboard_redirect(request):
    if request.user.is_superuser or request.user.rol == 'admin':
        return redirect(reverse('activos:admin_dashboard'))
    elif request.user.rol == 'logistica':
        return redirect(reverse('activos:logistica_dashboard'))
    elif request.user.rol == 'lectura':
        return redirect(reverse('activos:lectura_dashboard'))
    elif request.user.rol == 'asignador':
        return redirect(reverse('activos:home'))
    else:
        return redirect(reverse('activos:home'))

@login_required
def admin_dashboard(request):
    if not (request.user.is_superuser or request.user.rol == 'admin'):
        return redirect('activos:home')
    
    # Estadísticas generales
    total_activos = Activo.objects.count()
    asignados = Activo.objects.filter(estado__icontains='asignado').count()
    en_bodega = Activo.objects.filter(estado__icontains='confirmado').count()
    dados_baja = Activo.objects.filter(estado__icontains='baja').count()
    
    # Activos por zona
    from django.db.models import Count
    activos_por_zona = Activo.objects.values('zona').annotate(
        total=Count('item')
    ).order_by('-total')[:5]  # Top 5 zonas
    
    # Activos por categoría
    activos_por_categoria = Activo.objects.values('categoria__nombre').annotate(
        total=Count('item')
    ).order_by('-total')
    
    # Activos por estado
    activos_por_estado = Activo.objects.values('estado').annotate(
        total=Count('item')
    ).order_by('-total')
    
    # Últimos 5 activos registrados
    ultimos_activos = Activo.objects.order_by('-fecha_creacion')[:5]
    
    # Activos por operador
    activos_por_operador = Activo.objects.exclude(
        operador__isnull=True
    ).exclude(
        operador=''
    ).values('operador').annotate(
        total=Count('item')
    ).order_by('-total')
    
    return render(request, 'activos/admin_dashboard.html', {
        'total_activos': total_activos,
        'asignados': asignados,
        'en_bodega': en_bodega,
        'dados_baja': dados_baja,
        'activos_por_zona': activos_por_zona,
        'activos_por_categoria': activos_por_categoria,
        'activos_por_estado': activos_por_estado,
        'ultimos_activos': ultimos_activos,
        'activos_por_operador': activos_por_operador,
    })

@login_required
def logistica_dashboard(request):
    if request.user.rol != 'logistica':
        return redirect('activos:home')
    activos_por_estado = Activo.objects.values('estado').annotate(count=models.Count('estado'))
    movimientos_recientes = Tranzabilidad.objects.order_by('-fecha')[:10]
    return render(request, 'activos/logistica_dashboard.html', {
        'activos_por_estado': activos_por_estado,
        'movimientos_recientes': movimientos_recientes,
    })

@login_required
def lectura_dashboard(request):
    if request.user.rol != 'lectura':
        return redirect('activos:home')
    total_activos = Activo.objects.count()
    return render(request, 'activos/lectura_dashboard.html', {
        'total_activos': total_activos,
    })

class ActivoListView(LoginRequiredMixin, ListView):
    model = Activo
    template_name = 'activos/home.html'
    context_object_name = 'activos'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtro por documento del asesor
        documento = self.request.GET.get('documento', '').strip()
        if documento:
            queryset = queryset.filter(documento__icontains=documento)
        
        # Filtro por S/N
        sn = self.request.GET.get('sn', '').strip()
        if sn:
            queryset = queryset.filter(sn__icontains=sn)

        # Filtro por activo
        activo = self.request.GET.get('activo', '').strip()
        if activo:
            queryset = queryset.filter(activo__icontains=activo)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasar los valores de los filtros al contexto para mantenerlos en el formulario
        context['filtro_documento'] = self.request.GET.get('documento', '')
        context['filtro_sn'] = self.request.GET.get('sn', '')
        context['filtro_activo'] = self.request.GET.get('activo', '')
        return context


@login_required
def exportar_excel(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    activos = Activo.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Activos"

    # Definir estilos
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['ITEM', 'DOCUMENTO', 'NOMBRES Y APELLIDOS', 'IMEI 1', 'IMEI 2', 'S/N', 
               'ICCID', 'OPERADOR', 'MAC SUPERFLEX', 'MARCA', 'ACTIVO', 'CARGO', 'ESTADO', 
               'FECHA CONFIRMACIÓN', 'RESPONSABLE', 'IDENTIFICACIÓN', 'ZONA', 'CATEGORÍA', 
               'OBSERVACIÓN', 'PUNTO DE VENTA', 'CÓDIGO CENTRO COSTO', 'CENTRO COSTO PUNTO', 
               'FECHA SALIDA BODEGA']
    
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data
    for activo in activos:
        row = [
            activo.item,
            activo.documento or '',
            activo.nombres_apellidos or '',
            activo.imei1 or '',
            activo.imei2 or '',
            activo.sn or '',
            activo.iccid or '',
            activo.operador or '',
            activo.mac_superflex or '',
            activo.marca or '',
            activo.activo or '',
            activo.cargo or '',
            activo.estado or '',
            activo.fecha_confirmacion.strftime('%d/%m/%Y') if activo.fecha_confirmacion else '',
            activo.responsable or '',
            activo.identificacion or '',
            activo.zona or '',
            str(activo.categoria) if activo.categoria else '',
            activo.observacion or '',
            activo.punto_venta or '',
            activo.codigo_centro_costo or '',
            activo.centro_costo_punto or '',
            activo.fecha_salida_bodega.strftime('%d/%m/%Y') if activo.fecha_salida_bodega else ''
        ]
        ws.append(row)
        
        # Aplicar bordes a las celdas de datos
        for cell in ws[ws.max_row]:
            cell.border = border

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventario_activos.xlsx'
    wb.save(response)
    return response

class ActivoCreateView(LoginRequiredMixin, CreateView):
    model = Activo
    form_class = ActivoForm
    template_name = 'activos/activo_form.html'
    success_url = reverse_lazy('activos:home')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['is_update'] = False
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Datos para autocompletado
        context['documentos'] = Activo.objects.exclude(documento__isnull=True).exclude(documento='').values_list('documento', flat=True).distinct()
        context['nombres'] = Activo.objects.exclude(nombres_apellidos__isnull=True).exclude(nombres_apellidos='').values_list('nombres_apellidos', flat=True).distinct()
        context['identificaciones'] = Activo.objects.exclude(identificacion__isnull=True).exclude(identificacion='').values_list('identificacion', flat=True).distinct()
        context['codigos_centro'] = Activo.objects.exclude(codigo_centro_costo__isnull=True).exclude(codigo_centro_costo='').values_list('codigo_centro_costo', flat=True).distinct()
        context['nombres_centro'] = Activo.objects.exclude(centro_costo_punto__isnull=True).exclude(centro_costo_punto='').values_list('centro_costo_punto', flat=True).distinct()
        
        # Datos de marcas agrupadas por categoría para filtrado dinámico
        marcas_por_categoria = {}
        for marca in Marca.objects.select_related('categoria').all():
            cat_id = marca.categoria.id
            if cat_id not in marcas_por_categoria:
                marcas_por_categoria[cat_id] = []
            marcas_por_categoria[cat_id].append({'id': marca.id, 'nombre': marca.nombre})
        context['marcas_por_categoria_json'] = json.dumps(marcas_por_categoria)
        
        context['marcas_por_categoria_json'] = json.dumps(marcas_por_categoria)
        
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        Tranzabilidad.objects.create(
            activo=self.object,
            tipo='ingreso',
            usuario=self.request.user,
            zona_destino=self.object.zona,
            estado_nuevo=self.object.estado,
            descripcion='Ingreso inicial del activo al sistema'
        )
        return response

class ActivoDetailView(LoginRequiredMixin, DetailView):
    model = Activo
    template_name = 'activos/activo_detail.html'

class ActivoUpdateView(LoginRequiredMixin, UpdateView):
    model = Activo
    form_class = ActivoForm
    template_name = 'activos/activo_form.html'
    success_url = reverse_lazy('activos:home')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['is_update'] = True
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Datos para autocompletado
        context['documentos'] = Activo.objects.exclude(documento__isnull=True).exclude(documento='').values_list('documento', flat=True).distinct()
        context['nombres'] = Activo.objects.exclude(nombres_apellidos__isnull=True).exclude(nombres_apellidos='').values_list('nombres_apellidos', flat=True).distinct()
        context['identificaciones'] = Activo.objects.exclude(identificacion__isnull=True).exclude(identificacion='').values_list('identificacion', flat=True).distinct()
        context['codigos_centro'] = Activo.objects.exclude(codigo_centro_costo__isnull=True).exclude(codigo_centro_costo='').values_list('codigo_centro_costo', flat=True).distinct()
        context['nombres_centro'] = Activo.objects.exclude(centro_costo_punto__isnull=True).exclude(centro_costo_punto='').values_list('centro_costo_punto', flat=True).distinct()
        
        # Datos de marcas agrupadas por categoría para filtrado dinámico
        marcas_por_categoria = {}
        for marca in Marca.objects.select_related('categoria').all():
            cat_id = marca.categoria.id
            if cat_id not in marcas_por_categoria:
                marcas_por_categoria[cat_id] = []
            marcas_por_categoria[cat_id].append({'id': marca.id, 'nombre': marca.nombre})
        context['marcas_por_categoria_json'] = json.dumps(marcas_por_categoria)
        
        context['marcas_por_categoria_json'] = json.dumps(marcas_por_categoria)
        
        return context

    def form_valid(self, form):
        # Obtener estado anterior de la base de datos
        activo_anterior = Activo.objects.get(pk=self.object.pk)
        
        response = super().form_valid(form)
        
        # Determinar tipo de movimiento
        tipo = 'actualizacion'
        cambios = []
        
        # Campos a ignorar o tratar especial
        ignored_fields = ['fecha_actualizacion', 'usuario_actualizacion'] 
        
        for field in form.changed_data:
            if field in ignored_fields:
                continue
                
            old_value = getattr(activo_anterior, field)
            new_value = form.cleaned_data.get(field)
            
            # String representation for values
            if hasattr(old_value, 'nombre'): # For ForeignKey like Categoria/Marca
                old_str = old_value.nombre
            else:
                old_str = str(old_value) if old_value is not None else '-'
                
            if hasattr(new_value, 'nombre'):
                new_str = new_value.nombre
            else:
                new_str = str(new_value) if new_value is not None else '-'
            
            cambios.append(f"{field}: {old_str} -> {new_str}")
            
        descripcion = "Actualización: " + ", ".join(cambios) if cambios else "Actualización sin cambios detectados"
        
        if activo_anterior.estado != self.object.estado:
            tipo = 'cambio_estado'
            # Priorizar cambio de estado en la descripción pero mantener historial de otros campos
            descripcion = f'Cambio de estado: {activo_anterior.estado} -> {self.object.estado}. ' + descripcion
            
        tranzabilidad = Tranzabilidad.objects.create(
            activo=self.object,
            tipo=tipo,
            usuario=self.request.user,
            zona_origen=activo_anterior.zona,
            zona_destino=self.object.zona,
            estado_anterior=activo_anterior.estado,
            estado_nuevo=self.object.estado,
            descripcion=descripcion
        )

        # Si es una actualización normal (no cambio de estado), registrar campos cambiados
        if tipo == 'actualizacion' and form.changed_data:
            cambios = []
            for field in form.changed_data:
                # Omitir campos que no son relevantes para el historial visual o son internos
                if field in ['fecha_modificacion', 'usuario_modificacion']: 
                    continue
                
                try:
                    valor_ante = getattr(activo_anterior, field)
                    valor_nuev = form.cleaned_data.get(field)
                    
                    # Manejar campos ForeignKey (mostrar string representation)
                    if hasattr(valor_ante, 'nombre'):
                        valor_ante = valor_ante.nombre
                    if hasattr(valor_nuev, 'nombre'):
                        valor_nuev = valor_nuev.nombre
                        
                    # Manejar None/Empty
                    valor_ante = str(valor_ante) if valor_ante is not None else "Vacío"
                    valor_nuev = str(valor_nuev) if valor_nuev is not None else "Vacío"
                    
                    verbose_name = self.model._meta.get_field(field).verbose_name
                    cambios.append(f"{verbose_name}: {valor_ante} ➝ {valor_nuev}")
                except Exception as e:
                    continue
            
            if cambios:
                # Actualizar la descripción con los detalles
                tranzabilidad.descripcion = "Actualización de datos:\n" + "\n".join(cambios)
                tranzabilidad.save()
        return response

class ActivoAsignarView(LoginRequiredMixin, UpdateView):
    model = Activo
    form_class = ActivoForm
    template_name = 'activos/activo_form.html'
    success_url = reverse_lazy('activos:home')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['is_update'] = True
        kwargs['is_assignment'] = True
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Asignar Activo'
        # Datos para autocompletado (igual que en UpdateView)
        context['documentos'] = Activo.objects.exclude(documento__isnull=True).exclude(documento='').values_list('documento', flat=True).distinct()
        context['nombres'] = Activo.objects.exclude(nombres_apellidos__isnull=True).exclude(nombres_apellidos='').values_list('nombres_apellidos', flat=True).distinct()
        context['identificaciones'] = Activo.objects.exclude(identificacion__isnull=True).exclude(identificacion='').values_list('identificacion', flat=True).distinct()
        context['codigos_centro'] = Activo.objects.exclude(codigo_centro_costo__isnull=True).exclude(codigo_centro_costo='').values_list('codigo_centro_costo', flat=True).distinct()
        context['nombres_centro'] = Activo.objects.exclude(centro_costo_punto__isnull=True).exclude(centro_costo_punto='').values_list('centro_costo_punto', flat=True).distinct()
        
        # Datos de marcas agrupadas por categoría para filtrado dinámico
        marcas_por_categoria = {}
        for marca in Marca.objects.select_related('categoria').all():
            cat_id = marca.categoria.id
            if cat_id not in marcas_por_categoria:
                marcas_por_categoria[cat_id] = []
            marcas_por_categoria[cat_id].append({'id': marca.id, 'nombre': marca.nombre})
        context['marcas_por_categoria_json'] = json.dumps(marcas_por_categoria)
        
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        Tranzabilidad.objects.create(
            activo=self.object,
            tipo='asignacion',
            usuario=self.request.user,
            zona_destino=self.object.zona,
            descripcion=f'Asignado a: {self.object.responsable}'
        )
        return response

class ActivoDeleteView(LoginRequiredMixin, DeleteView):
    model = Activo
    template_name = 'activos/activo_confirm_delete.html'
    success_url = reverse_lazy('activos:home')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para eliminar activos.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        success_url = self.get_success_url()
        # Registrar trazabilidad antes de eliminar
        Tranzabilidad.objects.create(
            activo=self.object,
            tipo='eliminacion',
            usuario=self.request.user,
            descripcion=f'Activo eliminado. Serial: {self.object.sn}, Documento: {self.object.documento}'
        )
        self.object.delete()
        return HttpResponseRedirect(success_url)


@login_required
def eliminar_multiples_activos(request):
    """Vista para eliminar múltiples activos"""
    if request.user.rol != 'admin':
        messages.error(request, 'No tienes permisos para eliminar activos.')
        return redirect('activos:home')
    
    if request.method == 'POST':
        # Obtener los IDs de los activos a eliminar
        activos_ids = request.POST.getlist('activos_ids')
        
        if not activos_ids:
            messages.warning(request, 'No se seleccionaron activos para eliminar.')
            return redirect('activos:home')
        
        # Eliminar los activos
        activos_eliminados = Activo.objects.filter(pk__in=activos_ids)
        cantidad = activos_eliminados.count()
        
        # Registrar trazabilidad para cada activo antes de eliminar
        for activo in activos_eliminados:
            Tranzabilidad.objects.create(
                activo=activo,
                tipo='eliminacion',
                usuario=request.user,
                descripcion=f'Activo eliminado (Masivo). Serial: {activo.sn}, Documento: {activo.documento}'
            )

        activos_eliminados.delete()
        
        messages.success(request, f'Se eliminaron {cantidad} activo(s) exitosamente.')
        return redirect('activos:home')
    
    # GET request - mostrar confirmación
    activos_ids = request.GET.getlist('ids')
    
    if not activos_ids:
        messages.warning(request, 'No se seleccionaron activos para eliminar.')
        return redirect('activos:home')
    
    activos = Activo.objects.filter(pk__in=activos_ids)
    
    return render(request, 'activos/activo_confirm_delete_multiple.html', {
        'activos': activos,
        'activos_ids': activos_ids,
    })





# Tranzabilidad CRUD
class TranzabilidadListView(LoginRequiredMixin, ListView):
    model = Tranzabilidad
    template_name = 'activos/tranzabilidad_list.html'
    context_object_name = 'movimientos'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol not in ['admin', 'logistica', 'asignador']:
            messages.error(request, 'No tienes permisos para ver tranzabilidad.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('activo', 'usuario', 'activo__categoria', 'activo__marca')
        
        # Filtros
        tipo = self.request.GET.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        
        fecha_inicio = self.request.GET.get('fecha_inicio')
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        
        fecha_fin = self.request.GET.get('fecha_fin')
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)
        
        usuario = self.request.GET.get('usuario')
        if usuario:
            queryset = queryset.filter(usuario_id=usuario)
        
        activo_query = self.request.GET.get('activo')
        if activo_query:
            queryset = queryset.filter(
                Q(activo__activo__icontains=activo_query) |
                Q(activo__sn__icontains=activo_query) |
                Q(activo__item__icontains=activo_query)
            )
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Para los filtros
        context['tipos_accion'] = Tranzabilidad.TIPO_CHOICES
        context['usuarios'] = get_user_model().objects.filter(
            id__in=Tranzabilidad.objects.values_list('usuario_id', flat=True).distinct()
        )
        
        # Mantener valores de filtros
        context['filtro_tipo'] = self.request.GET.get('tipo', '')
        context['filtro_fecha_inicio'] = self.request.GET.get('fecha_inicio', '')
        context['filtro_fecha_fin'] = self.request.GET.get('fecha_fin', '')
        context['filtro_usuario'] = self.request.GET.get('usuario', '')
        context['filtro_activo'] = self.request.GET.get('activo', '')
        
        # Historial de cambios por movimiento (para el modal)
        # Optimizacion: Evitar consultas N+1 cargando historiales relevantes en memoria o filtrando
        # Dado que es paginado, podemos hacer la consulta para los items de la pagina actual
        movimientos_page = context['movimientos'] # Esto es la page_obj o lista paginada en ListView
        
        historiales = {}
        for mov in movimientos_page:
            if mov.activo and mov.tipo in ['actualizacion', 'cambio_estado']:
                # Buscamos historiales cercanos a la fecha del movimiento
                # Usamos un rango pequeño de tiempo porque el historial se crea en la misma transacción/flujo
                historiales[mov.id] = Historial.objects.filter(
                    activo=mov.activo,
                    fecha__range=[mov.fecha - timedelta(seconds=2), mov.fecha + timedelta(seconds=2)]
                ).order_by('fecha')
        
        context['historiales'] = historiales
        
        return context

# Tranzabilidad registration
class RegistrarTranzabilidadView(LoginRequiredMixin, CreateView):
    model = Tranzabilidad
    fields = ['tipo', 'zona_origen', 'zona_destino', 'estado_nuevo', 'descripcion']
    template_name = 'activos/registrar_tranzabilidad.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol not in ['admin', 'logistica']:
            messages.error(request, 'No tienes permisos para registrar tranzabilidad.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activo'] = get_object_or_404(Activo, pk=self.kwargs['pk'])
        # Obtener zonas únicas de los activos
        context['zonas'] = Activo.objects.values_list('zona', flat=True).distinct()
        return context

    def form_valid(self, form):
        activo = get_object_or_404(Activo, pk=self.kwargs['pk'])
        form.instance.activo = activo
        form.instance.usuario = self.request.user
        form.instance.estado_anterior = activo.estado

        if form.cleaned_data['estado_nuevo']:
            activo.estado = form.cleaned_data['estado_nuevo']
            activo.save()

        if form.cleaned_data['zona_destino']:
            activo.zona = form.cleaned_data['zona_destino']
            activo.save()

        messages.success(self.request, 'Tranzabilidad registrada exitosamente.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('activos:activo_detail', kwargs={'pk': self.kwargs['pk']})


# Historial de activo
@login_required
def historial_activo(request, pk):
    activo = get_object_or_404(Activo, pk=pk)
    historial = Historial.objects.filter(activo=activo).order_by('-fecha')
    movimientos = Tranzabilidad.objects.filter(activo=activo).order_by('-fecha')
    return render(request, 'activos/historial_activo.html', {
        'activo': activo,
        'historial': historial,
        'movimientos': movimientos,
    })

# Reporte por sede
@login_required
def reporte_por_sede(request):
    if request.user.rol not in ['admin', 'logistica', 'asignador']:
        messages.error(request, 'No tienes permisos para ver reportes.')
        return redirect('activos:home')

    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    zona = request.GET.get('zona')
    categoria = request.GET.get('categoria')
    estado = request.GET.get('estado')

    # Base queryset
    activos = Activo.objects.all().select_related('categoria')

    # Aplicar filtros
    if fecha_inicio:
        activos = activos.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        activos = activos.filter(fecha_creacion__date__lte=fecha_fin)
    if zona:
        activos = activos.filter(zona=zona)
    if categoria:
        activos = activos.filter(categoria_id=categoria)
    if estado:
        activos = activos.filter(estado=estado)

    # Datos para los filtros
    zonas = Activo.objects.values_list('zona', flat=True).distinct().order_by('zona')
    categorias = Categoria.objects.all().order_by('nombre')
    estados = Activo.objects.values_list('estado', flat=True).distinct().order_by('estado')

    return render(request, 'activos/reporte_por_sede.html', {
        'activos': activos,
        'zonas': zonas,
        'categorias': categorias,
        'estados': estados,
        # Mantener filtros seleccionados
        'filtro_fecha_inicio': fecha_inicio,
        'filtro_fecha_fin': fecha_fin,
        'filtro_zona': zona,
        'filtro_categoria': int(categoria) if categoria else '',
        'filtro_estado': estado,
    })

@login_required
def exportar_excel(request):
    if request.user.rol not in ['admin', 'logistica', 'asignador']:
        messages.error(request, 'No tienes permisos para exportar reportes.')
        return redirect('activos:home')

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Obtener parámetros de filtro (mismos que en reporte)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    zona = request.GET.get('zona')
    categoria = request.GET.get('categoria')
    estado = request.GET.get('estado')

    # Base queryset
    activos = Activo.objects.all().select_related('categoria')

    # Aplicar filtros
    if fecha_inicio:
        activos = activos.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        activos = activos.filter(fecha_creacion__date__lte=fecha_fin)
    if zona:
        activos = activos.filter(zona=zona)
    if categoria:
        activos = activos.filter(categoria_id=categoria)
    if estado:
        activos = activos.filter(estado=estado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Activos"

    # Definir estilos
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers completos
    headers = ['ITEM', 'DOCUMENTO', 'NOMBRES Y APELLIDOS', 'IMEI 1', 'IMEI 2', 'S/N', 
               'ICCID', 'OPERADOR', 'MAC SUPERFLEX', 'MARCA', 'ACTIVO', 'CARGO', 'ESTADO', 
               'FECHA CONFIRMACIÓN', 'RESPONSABLE', 'IDENTIFICACIÓN', 'ZONA', 'CATEGORÍA', 
               'OBSERVACIÓN', 'PUNTO DE VENTA', 'CÓDIGO CENTRO COSTO', 'CENTRO COSTO PUNTO', 
               'FECHA SALIDA BODEGA', 'FECHA CREACIÓN']
    
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data
    for activo in activos:
        row = [
            activo.item,
            activo.documento or '',
            activo.nombres_apellidos or '',
            activo.imei1 or '',
            activo.imei2 or '',
            activo.sn or '',
            activo.iccid or '',
            activo.operador or '',
            activo.mac_superflex or '',
            str(activo.marca) if activo.marca else '',
            activo.activo or '',
            activo.cargo or '',
            activo.estado or '',
            activo.fecha_confirmacion.strftime('%d/%m/%Y') if activo.fecha_confirmacion else '',
            activo.responsable or '',
            activo.identificacion or '',
            activo.zona or '',
            str(activo.categoria) if activo.categoria else '',
            activo.observacion or '',
            activo.punto_venta or '',
            activo.codigo_centro_costo or '',
            activo.centro_costo_punto or '',
            activo.fecha_salida_bodega.strftime('%d/%m/%Y') if activo.fecha_salida_bodega else '',
            activo.fecha_creacion.strftime('%d/%m/%Y %H:%M') if activo.fecha_creacion else ''
        ]
        ws.append(row)
        
        # Aplicar bordes
        for cell in ws[ws.max_row]:
            cell.border = border

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_activos.xlsx'
    wb.save(response)
    return response







# Zona CRUD

class ZonaListView(LoginRequiredMixin, ListView):
    model = Zona
    template_name = 'activos/zona_list.html'
    context_object_name = 'zonas'

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para ver zonas.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

class ZonaCreateView(LoginRequiredMixin, CreateView):
    model = Zona
    fields = '__all__'
    template_name = 'activos/zona_form.html'
    success_url = reverse_lazy('activos:zona_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para crear zonas.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

class ZonaUpdateView(LoginRequiredMixin, UpdateView):
    model = Zona
    fields = '__all__'
    template_name = 'activos/zona_form.html'
    success_url = reverse_lazy('activos:zona_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para editar zonas.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

class ZonaDeleteView(LoginRequiredMixin, DeleteView):
    model = Zona
    template_name = 'activos/zona_confirm_delete.html'
    success_url = reverse_lazy('activos:zona_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para eliminar zonas.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)


# Categoria CRUD

class CategoriaListView(LoginRequiredMixin, ListView):
    model = Categoria
    template_name = 'activos/category_list.html'
    context_object_name = 'categorias'

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para ver categorías.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

class CategoriaCreateView(LoginRequiredMixin, CreateView):
    model = Categoria
    fields = ['nombre']
    template_name = 'activos/category_form.html'
    success_url = reverse_lazy('activos:categoria_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para crear categorías.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        # Guardar la categoría primero
        response = super().form_valid(form)
        
        # Procesar las marcas
        marca_nombres = self.request.POST.getlist('marca_nombre[]')
        
        for nombre in marca_nombres:
            if nombre.strip():  # Solo crear si el nombre no está vacío
                Marca.objects.create(
                    nombre=nombre.strip(),
                    categoria=self.object
                )
        
        messages.success(self.request, f'Categoría "{self.object.nombre}" creada exitosamente con {len([n for n in marca_nombres if n.strip()])} marca(s).')
        return response

class CategoriaUpdateView(LoginRequiredMixin, UpdateView):
    model = Categoria
    fields = ['nombre']
    template_name = 'activos/category_form.html'
    success_url = reverse_lazy('activos:categoria_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para editar categorías.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        # Guardar la categoría primero
        response = super().form_valid(form)
        
        # Obtener marcas del formulario
        marca_nombres = self.request.POST.getlist('marca_nombre[]')
        marca_ids = self.request.POST.getlist('marca_id[]')
        
        # IDs de marcas que se mantienen o actualizan
        marcas_actualizadas = []
        
        # Procesar marcas
        for i, nombre in enumerate(marca_nombres):
            if nombre.strip():
                marca_id = marca_ids[i] if i < len(marca_ids) and marca_ids[i] else None
                
                if marca_id:  # Marca existente - actualizar
                    try:
                        marca = Marca.objects.get(pk=marca_id, categoria=self.object)
                        marca.nombre = nombre.strip()
                        marca.save()
                        marcas_actualizadas.append(int(marca_id))
                    except Marca.DoesNotExist:
                        pass
                else:  # Marca nueva - crear
                    nueva_marca = Marca.objects.create(
                        nombre=nombre.strip(),
                        categoria=self.object
                    )
                    marcas_actualizadas.append(nueva_marca.id)
        
        # Eliminar marcas que ya no están en el formulario
        Marca.objects.filter(categoria=self.object).exclude(id__in=marcas_actualizadas).delete()
        
        messages.success(self.request, f'Categoría "{self.object.nombre}" actualizada exitosamente.')
        return response

class CategoriaDeleteView(LoginRequiredMixin, DeleteView):
    model = Categoria
    template_name = 'activos/category_confirm_delete.html'
    success_url = reverse_lazy('activos:categoria_list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol != 'admin':
            messages.error(request, 'No tienes permisos para eliminar categorías.')
            return redirect('activos:home')
        return super().dispatch(request, *args, **kwargs)

